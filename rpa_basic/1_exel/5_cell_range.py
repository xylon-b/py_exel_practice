from openpyxl import Workbook
from random import *
wb = Workbook()
ws = wb.active

#1줄씩 데이터 넣기
ws.append(["번호", "영어", "수학"])
for i in range(1, 11): #10개 데이터 넣기
    ws.append([i, randint(0, 100), randint(0, 100)])

col_B = ws["B"] #영어 컬럼만 가지고 오기
# #print(col_B)
# for cell in col_B:
#     print(cell.value)

col_range = ws["B:C"] #영어, 수락 column 함꼐 가지고 오기
# for cols in col_range:
    # for cell in cols:
        # print(cell.value)

row_title = ws[1] #첫번째 row만 가지고 오기
#for cell in row_title:
#    print(cell.value)

#row_range = ws[2:6]
#for rows in row_range:
#    for cell in rows:
#        print(cell.value, end=" ")
#    print()

from openpyxl.utils.cell import column_index_from_string, coordinate_from_string

row_range = ws[2:ws.max_row]
#for rows in row_range:
#    for cell in rows:
#        #print(cell.coordinate, end=" ")
#        xy = coordinate_from_string(cell.coordinate)
#        #print(xy, end=" ")
#        print(xy[0], end="") # A
#        print(xy[1], end=" ") # 1
#    print()

#전체 rows
#print(tuple(ws.rows))
# for row in tuple(ws.rows):
#     print(row[2].value)

#전제 columns
#print(tuple(ws.columns))

# for columns in tuple(ws.columns):
#     print(columns[0].value)

#for row in ws.iter_rows(): # 전체 row
    # print(row[1].value)

# 2번째 줄부터 11번째 줄까지, 2번째 열부터 3번쨰 열까지
for row in ws.iter_rows(min_row=2, max_row=11, min_col=2, max_col=3): # 전체 row
    #print(row[0].value, row[1].value)
    print(row)

for col in ws.iter_cols(min_row=1, max_row=5, min_col=1, max_col=3):
    print(col)



#for column in ws.iter_cols():
#    print(column[0].value)

wb.save("sample.xlsx")