from openpyxl import Workbook
wb = Workbook()
ws = wb.active

#데이터를 넣기
ws.append(("학번", "출석", "퀴즈1", "퀴즈2", "중간고사", "기말고사", "프로젝트"))


# 기존 성적 업데이트

# 퀴즈2 점수를 10으로 수정


# H열에 총점 (sum이용), I열에 성적정보 추가

ws["H1"] = "총점"
ws["I1"] = "성적"

# idx를 2부터 시작하게

    #총점 90이상

    # 출석이 5점 미만이면 F
  

    # I열에 성적 정보 추가
   
wb.save("scores.xlsx")
