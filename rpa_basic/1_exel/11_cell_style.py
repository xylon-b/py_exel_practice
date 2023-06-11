from openpyxl import styles
from openpyxl.styles import Font, Border, PatternFill, Alignment
from openpyxl import load_workbook
from openpyxl.styles.borders import Side
wb = load_workbook("sample.xlsx")
ws = wb.active

#번호 영어 수학
a1 = ws["A1"]
b1 = ws["B1"]
c1 = ws["C1"]

# A 열의 너미를 5로
ws.column_dimensions["A"].width = 5

#1행의 높이를 50으로 설정
ws.row_dimensions[1].height = 50

#스타일 적용
a1.font = Font(color="FF0000", italic=True, bold=True)
b1.font = Font(color="CC33FF", name="Arial", strike=True) #폰트를 Arial 로 설정, 쉬소선
c1.font = Font(color="0000FF", size=20, underline="single") # 글자크리 20 밑줄적용

# 테두리 적용
thin_border = Border(left=Side(style="thin"), right=Side(style="thin"), bottom=Side(style="thin"))
a1.border = thin_border
b1.border = thin_border
c1.border = thin_border

# 90 점 넘는 셀에 대해서 초록새긍로 적용
for row in ws.rows:
    for cell in row:
        #각 셀에 대해서 정렬
        cell.alignment = Alignment(horizontal="center", vertical="center")
        #center, left, right, top, bottom
        if cell.column == 1:
            continue

        # cell 이 정수형 데이터이고 90점보다 놓으면
        if isinstance(cell.value, int) and cell.value > 90:
            cell.fill = PatternFill(fgColor="00FF00", fill_type="solid") #배경을 초록색으로
            cell.font = Font(color="FF0000") #폰트 색상 변경

#틀 고정
ws.freeze_panes = "B2" #B2 기준으로 틀 고정


wb.save("sample_style.xlsx")
