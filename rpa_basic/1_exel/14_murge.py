from openpyxl import Workbook
wb = Workbook()
ws = wb.active

# 병합하기
ws.merge_cells("B2:D2")
ws["B2"].value = "Merged Cell"
print(ws["B2"].value)

wb.save("sample_merge.xlsx")